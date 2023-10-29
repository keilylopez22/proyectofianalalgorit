from BD.conexionconsol import DAOconsol
import datetime
import openpyxl
import csv

def listar_ventas(ventas):
    print("\nVENTAS: \n")
    contador= 1
    for venta in ventas:
        datos = "{0}. CODIGO_DE_VENTA: {1} | CODIGO_CLIENTE: {2}  |CODIGO_PRODUCTOS: {3} |CANTIDAD_DE_PRODUCTOS:{4},|TOTAL_DE_VENTA: {5} ({6} FECHA)"
        print(datos.format(contador, venta[0], venta[1], venta[2], venta[3], venta[4],venta[5]))
        contador= contador+1
        print(" ")




#esta funcion es la que nos va a guardar los datos de las ventas que creermos
def pedirdatosventas():
     fecha_venta = datetime.datetime.now()
       #codigo_correcto= False
       #while(not codigo_correcto):
       #codigo_producto = int(input("ingrese el codigo: "))
          #if len(codigo_producto) <=6:
           # codigo_correcto= True
          #else:
              # print("codigo incorrecto: el codigo no debe ser mayor a 6 digitos")    
       
       
     codigo_cliente_correcto = False
     while(not codigo_cliente_correcto):
          codigo_cliente= input("ingrese el codigo del cliente para generar la venta: ")
     
          if codigo_cliente.isnumeric():
               if (int(codigo_cliente) > 0):
                    codigo_cliente_correcto= True
                    codigo_cliente= int(codigo_cliente)
               else:
                    print("lo siento, el codigo que generaste no puede ser un numero menor o igual a 0")  
          else:
               print("codigo cliente incorrecto: parece que ingresaste un texto, ingtresa un numero por favor")     

     codigo_producto_correcto= False
     while( not codigo_producto_correcto):
          
          codigo_productos= (input("ingresa el codigo del producto para generar la venta: "))
          if codigo_productos.isnumeric():
               if (int(codigo_productos) > 0):
                    codigo_producto_correcto= True
                    codigo_productos= int(codigo_productos)
                    
               else:
                    print("el codigo debe ser mayor a 0")
          else:
               print("codigo del producto incorrecto: parece que ingresaste un texto, ingresa valores numericos por favor")

     cantidad_prod_correcta= False
     while( not cantidad_prod_correcta):
          
          cantidad_productos= (input("ingresa la cantidad del los  productos para generar la venta: "))
          if cantidad_productos.isnumeric():
               if (int(cantidad_productos) > 0):
                    cantidad_prod_correcta= True
                    cantidad_productos= int(cantidad_productos)
                    
               else:
                    print("lo siento la cantidad debe ser mayor a 0 para que se genere una venta")
          else:
               print("codigo del producto incorrecto: parece que ingresaste un texto, ingresa valores numericos por favor")

     totalvencorrecta= False
     while( not totalvencorrecta):
          
          total_venta= (input("ingresa el monto total para generar la venta: "))
          if total_venta.isnumeric():
               if (float(total_venta) > 0):
                    totalvencorrecta= True
                    total_venta= int(total_venta)
                    
               else:
                    print("lo siento la cantidad debe ser mayor a 0 para que se genere una venta")
          else:
               print("parece que ingresaste un texto, ingresa valores numericos por favor")
     return  codigo_cliente, codigo_productos, cantidad_productos, total_venta, fecha_venta

     


def pedir_ventas_eliminacion(ventas):
     listar_ventas(ventas)
     codigoventaExiste= False
     ventaEliminar= input("ingrese el codigo de venta que deseaa eliminar: ")
     for venta in ventas:
          #print(str(venta[0]) +" = "+ ventaEliminar+" es: " + str(str(venta[0]) == str(ventaEliminar)))
          if (str (venta[0]) == str(ventaEliminar)):
               codigoventaExiste= True
               break
     if not codigoventaExiste:
          ventaEliminar = ""   
     return ventaEliminar





#para hacer el excel
def generar_excel(nombreReporte, ventas):
     
     

     print("openpyxl version:", openpyxl.__version__)
     """wb = openpyxl.Workbook()
     wb.create_sheet()
     wb.save(nombreReporte+".xlsx")
     wb.close()
     wb = openpyxl.load_workbook(nombreReporte+".xlsx")
     hoja = wb.create_sheet('hoja1')
     hoja['A1'] = "prueba"
     for venta in ventas:
          pass
     wb.save(nombreReporte)
     wb.close()"""

     file = open(nombreReporte + ".csv", "w")
     for venta in ventas:
          file.write(str(venta[0]) + "," + str(venta[1]) + "," + str(venta[2]) + "," + str(venta[3]) + "," + str(venta[4]) + "," + str(venta[5]) + "," + str(venta[6])+ "\n")
     file.close()

     